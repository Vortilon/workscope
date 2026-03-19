from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from rapidfuzz import fuzz


CANONICAL_COLUMNS: dict[str, list[str]] = {
    "task_reference": ["task", "task ref", "mpd task", "task no", "reference", "doc/rev", "task doc/rev", "item no"],
    "service_interval": ["service n", "interval", "check", "service", "frequency"],
    "description": ["description", "task description", "work description", "scope", "action"],
    "man_hours": ["mh", "man hours", "man hrs", "hours", "labour hours", "est. mh"],
    "ata_chapter": ["ata", "ata chapter", "ata ref", "chapter"],
    "status": ["status", "done", "completed", "sign off", "insp date", "reviewed by"],
    "component_pn": ["p/n", "part no", "part number", "pn"],
    "component_sn": ["s/n", "serial no", "serial number", "sn"],
    "part_type": ["type", "category", "class"],
    "unit": ["unit", "unid", "uom"],
    "quantity": ["qty", "quantity"],
}


@dataclass
class ExcelParseResult:
    header: dict[str, Any]
    sections: list[dict[str, Any]]
    parts_list: list[dict[str, Any]]
    parse_warnings: list[str]
    confidence: str
    totals: dict[str, Any]


def _best_match(header: str, candidates: list[str]) -> int:
    header_l = header.strip().lower()
    best = 0
    for c in candidates:
        best = max(best, fuzz.partial_ratio(header_l, c))
    return best


def _detect_header_row(rows: list[list[Any]]) -> int | None:
    # Find row with multiple recognizable headers.
    best_idx = None
    best_score = 0
    for i, row in enumerate(rows[:60]):
        texts = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
        if len(texts) < 3:
            continue
        score = 0
        for t in texts:
            score += max(
                _best_match(t, CANONICAL_COLUMNS["task_reference"]),
                _best_match(t, CANONICAL_COLUMNS["description"]),
                _best_match(t, CANONICAL_COLUMNS["man_hours"]),
            )
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score >= 240 else None


def _map_columns(headers: list[str]) -> tuple[dict[str, int], list[str]]:
    mapping: dict[str, int] = {}
    warnings: list[str] = []
    for canonical, candidates in CANONICAL_COLUMNS.items():
        best_j = None
        best_score = 0
        for j, h in enumerate(headers):
            score = _best_match(h, candidates)
            if score > best_score:
                best_score = score
                best_j = j
        if best_j is not None and best_score >= 75:
            mapping[canonical] = best_j
    if "task_reference" not in mapping and "description" not in mapping:
        warnings.append("Excel: could not confidently map task_reference/description columns")
    return mapping, warnings


def _unmerge_and_fill(ws) -> None:
    # openpyxl only supports merged cell ranges in non-readonly mode.
    merged = list(getattr(ws, "merged_cells", []).ranges)
    for r in merged:
        min_row, min_col, max_row, max_col = r.min_row, r.min_col, r.max_row, r.max_col
        v = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(r))
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                ws.cell(rr, cc).value = v


def parse_excel_generic(path: Path) -> ExcelParseResult:
    import openpyxl

    warnings: list[str] = []
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    sections: list[dict[str, Any]] = []
    parts_list: list[dict[str, Any]] = []
    all_tasks: list[dict[str, Any]] = []

    for sheet_index, ws in enumerate(wb.worksheets):
        # Skip very hidden sheets
        if getattr(ws, "sheet_state", "") == "hidden":
            continue
        _unmerge_and_fill(ws)

        # Build rows while skipping hidden rows/cols
        hidden_cols = {i for i, dim in ws.column_dimensions.items() if getattr(dim, "hidden", False)}
        hidden_rows = {i for i, dim in ws.row_dimensions.items() if getattr(dim, "hidden", False)}
        rows: list[list[Any]] = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if r_idx in hidden_rows:
                continue
            values: list[Any] = []
            for c_idx, cell in enumerate(row, start=1):
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                if col_letter in hidden_cols:
                    continue
                values.append(cell.value)
            rows.append(values)

        if not rows:
            continue

        header_row_idx = _detect_header_row(rows)
        if header_row_idx is None:
            warnings.append(f"Excel: sheet '{ws.title}' no header row detected; skipped")
            continue

        headers = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
        mapping, map_warn = _map_columns(headers)
        warnings.extend([f"Excel: sheet '{ws.title}': {w}" for w in map_warn])

        tasks: list[dict[str, Any]] = []
        for i, row in enumerate(rows[header_row_idx + 1 :], start=1):
            # Empty row skip
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue

            def get(col: str) -> Any:
                j = mapping.get(col)
                if j is None or j >= len(row):
                    return None
                return row[j]

            task_ref = get("task_reference")
            desc = get("description")
            if task_ref is None and desc is None:
                continue

            raw = {headers[j] if j < len(headers) else f"Col{j}": row[j] for j in range(min(len(headers), len(row)))}
            extra_fields = {}
            for j, h in enumerate(headers):
                if h == "":
                    continue
                if j in mapping.values():
                    continue
                if j < len(row) and row[j] is not None and str(row[j]).strip() != "":
                    extra_fields[h] = row[j]

            tasks.append(
                {
                    "line_number": len(all_tasks) + len(tasks) + 1,
                    "service_interval": str(get("service_interval")).strip() if get("service_interval") is not None else None,
                    "task_reference": str(task_ref).strip() if task_ref is not None else None,
                    "description": str(desc).strip() if desc is not None else None,
                    "man_hours": float(get("man_hours")) if isinstance(get("man_hours"), (int, float)) else None,
                    "status": str(get("status")).strip() if get("status") is not None else None,
                    "component_pn": str(get("component_pn")).strip() if get("component_pn") is not None else None,
                    "component_sn": str(get("component_sn")).strip() if get("component_sn") is not None else None,
                    "ata_chapter": str(get("ata_chapter")).strip() if get("ata_chapter") is not None else None,
                    "ata_derived": False if get("ata_chapter") is not None else True,
                    "raw_line": None,
                    "extra_fields": extra_fields,
                    "_sheet": ws.title,
                    "_sheet_index": sheet_index,
                    "_row_index": header_row_idx + 1 + i,
                    "_raw_row": raw,
                }
            )

        if tasks:
            all_tasks.extend(tasks)
            sections.append(
                {
                    "section_type": "UNKNOWN",
                    "task_count": len(tasks),
                    "tasks": tasks,
                    "extra_fields": {"sheet": ws.title},
                }
            )

    wb.close()

    totals = {
        "total_task_rows_in_document": len(all_tasks),
        "total_rows_extracted": len(all_tasks),
        "extraction_match": True,
        "total_mh_sum": float(sum(t["man_hours"] for t in all_tasks if isinstance(t.get("man_hours"), (int, float)))),
        "sections_found": [s["section_type"] for s in sections] if sections else [],
    }

    confidence = "medium" if sections else "low"
    return ExcelParseResult(
        header={"raw_fields": {}},
        sections=sections,
        parts_list=parts_list,
        parse_warnings=warnings,
        confidence=confidence,
        totals=totals,
    )

"""Web UI: Jinja + HTMX + Alpine. DAE styling with login gating."""
import json
import re as _re
import shutil
from datetime import datetime
from io import BytesIO
from pathlib import Path
from fastapi import APIRouter, Request, Depends, HTTPException, Form
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from app.config import MPD_STORAGE
from app.database import get_db
from app.models.mpd import MPDDataset, MPDTask
from app.models.operator import Operator
from app.models.aircraft_type import AircraftType, EngineType
from app.models.project import Project, ProjectConditionAnswer, ConditionAnswerHistory
from app.models.workscope import WorkscopeImportRow
from app.services.normalize import (
    normalize_applicability_tokens,
    threshold_tokens as _thr_tokens_fn,
    interval_tokens as _int_tokens_fn,
)

router = APIRouter()
BASE = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE / "templates"))

# ── Custom Jinja2 filters ────────────────────────────────────────────────────
templates.env.filters["threshold_tokens"] = _thr_tokens_fn
templates.env.filters["interval_tokens"]  = _int_tokens_fn


def _redirect_login():
    return RedirectResponse("/login", status_code=303)


def _require_login(request: Request):
    """Return RedirectResponse to login if not authenticated; else None."""
    if not request.session.get("user_id"):
        return _redirect_login()
    return None


def _require_admin(request: Request):
    """Return RedirectResponse to home if not admin; else None. Use after _require_login."""
    if request.session.get("role") != "admin":
        return RedirectResponse("/", status_code=303)
    return None


@router.get("/", response_class=HTMLResponse)
async def home(request: Request):
    if (r := _require_login(request)):
        return r
    return templates.TemplateResponse(request, "home.html")


@router.get("/mpd", response_class=HTMLResponse)
async def mpd_library(request: Request, error: str = "", db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    result = await db.execute(
        select(MPDDataset).order_by(MPDDataset.manufacturer, MPDDataset.model, MPDDataset.revision)
    )
    datasets = result.scalars().all()

    # Task counts per dataset (single query)
    counts_res = await db.execute(
        select(MPDTask.dataset_id, func.count().label("cnt")).group_by(MPDTask.dataset_id)
    )
    task_counts = {row.dataset_id: row.cnt for row in counts_res}

    # Which dataset IDs are referenced by at least one project
    used_res = await db.execute(
        select(Project.mpd_dataset_id).where(Project.mpd_dataset_id.is_not(None)).distinct()
    )
    in_use_ids = {row.mpd_dataset_id for row in used_res}

    # Which dataset IDs have a stored source file
    has_file_ids = set()
    for ds in datasets:
        for ext in (".xlsx", ".xls"):
            if (MPD_STORAGE / f"mpd_{ds.id}{ext}").exists():
                has_file_ids.add(ds.id)
                break

    return templates.TemplateResponse(request, "mpd_library.html", {
        "datasets": datasets,
        "task_counts": task_counts,
        "in_use_ids": in_use_ids,
        "has_file_ids": has_file_ids,
        "error": error,
    })


@router.get("/mpd/{dataset_id}", response_class=HTMLResponse)
async def mpd_detail(request: Request, dataset_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    ds_result = await db.execute(select(MPDDataset).where(MPDDataset.id == dataset_id))
    dataset = ds_result.scalars().one_or_none()
    if not dataset:
        raise HTTPException(404, "MPD dataset not found")
    tasks_result = await db.execute(
        select(MPDTask)
        .where(MPDTask.dataset_id == dataset_id)
        .order_by(MPDTask.row_index)
        .limit(1000)
    )
    tasks = tasks_result.scalars().all()
    return templates.TemplateResponse(request, "mpd_detail.html", {"dataset": dataset, "tasks": tasks})


@router.get("/projects", response_class=HTMLResponse)
async def project_list(request: Request, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    result = await db.execute(select(Project).order_by(Project.updated_at.desc()))
    projects = result.scalars().all()
    # Resolve operator names
    op_ids = {p.operator_id for p in projects if p.operator_id}
    op_map: dict[int, str] = {}
    if op_ids:
        ops_res = await db.execute(select(Operator).where(Operator.id.in_(op_ids)))
        op_map = {o.id: o.name for o in ops_res.scalars().all()}
    # Resolve aircraft type labels
    at_ids = {p.aircraft_type_id for p in projects if p.aircraft_type_id}
    at_map: dict[int, str] = {}
    if at_ids:
        at_res = await db.execute(select(AircraftType).where(AircraftType.id.in_(at_ids)))
        at_map = {a.id: a.display_name_with_series for a in at_res.scalars().all()}
    # Resolve engine type labels
    et_ids = {p.engine_type_id for p in projects if p.engine_type_id}
    et_map: dict[int, str] = {}
    if et_ids:
        et_res = await db.execute(select(EngineType).where(EngineType.id.in_(et_ids)))
        et_map = {e.id: e.engine_model for e in et_res.scalars().all()}
    return templates.TemplateResponse(request, "projects.html", {
        "projects": projects,
        "op_map": op_map,
        "at_map": at_map,
        "et_map": et_map,
    })


@router.get("/projects/new", response_class=HTMLResponse)
async def project_new(request: Request, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    datasets = (await db.execute(select(MPDDataset).order_by(MPDDataset.manufacturer))).scalars().all()
    operators = (await db.execute(select(Operator).order_by(Operator.name))).scalars().all()
    aircraft_types = (await db.execute(
        select(AircraftType).order_by(AircraftType.manufacturer, AircraftType.model)
    )).scalars().all()
    engine_types = (await db.execute(
        select(EngineType).order_by(EngineType.engine_manufacturer, EngineType.engine_family, EngineType.engine_model)
    )).scalars().all()
    return templates.TemplateResponse(request, "project_new.html", {
        "datasets": datasets,
        "operators": operators,
        "aircraft_types": aircraft_types,
        "engine_types": engine_types,
    })


@router.get("/projects/{project_id}", response_class=HTMLResponse)
async def project_detail(request: Request, project_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalars().one_or_none()
    if not project:
        raise HTTPException(404, "Project not found")
    operator_name = ""
    if project.operator_id:
        op = (await db.execute(select(Operator).where(Operator.id == project.operator_id))).scalars().one_or_none()
        if op:
            operator_name = op.name
    aircraft_type = None
    if project.aircraft_type_id:
        aircraft_type = (await db.execute(
            select(AircraftType).where(AircraftType.id == project.aircraft_type_id)
        )).scalars().one_or_none()
    engine_type = None
    if project.engine_type_id:
        engine_type = (await db.execute(
            select(EngineType).where(EngineType.id == project.engine_type_id)
        )).scalars().one_or_none()
    dataset = None
    if project.mpd_dataset_id:
        dataset = (await db.execute(select(MPDDataset).where(MPDDataset.id == project.mpd_dataset_id))).scalars().one_or_none()
    # Workscope import rows
    ws_rows_res = await db.execute(
        select(WorkscopeImportRow).where(WorkscopeImportRow.project_id == project_id).order_by(WorkscopeImportRow.seq)
    )
    ws_rows = ws_rows_res.scalars().all()
    return templates.TemplateResponse(request, "project_detail.html", {
        "project": project,
        "operator_name": operator_name,
        "aircraft_type": aircraft_type,
        "engine_type": engine_type,
        "dataset": dataset,
        "ws_rows": ws_rows,
    })


@router.get("/mpd/{dataset_id}/download")
async def mpd_dataset_download(request: Request, dataset_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    for ext in (".xlsx", ".xls"):
        fp = MPD_STORAGE / f"mpd_{dataset_id}{ext}"
        if fp.exists():
            ds = (await db.execute(select(MPDDataset).where(MPDDataset.id == dataset_id))).scalars().one_or_none()
            filename = (ds.source_file if ds and ds.source_file else f"mpd_{dataset_id}{ext}")
            return FileResponse(
                str(fp), filename=filename,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    raise HTTPException(404, "Source file not available")


@router.get("/mpd/{dataset_id}/export")
async def mpd_dataset_export(request: Request, dataset_id: int, db: AsyncSession = Depends(get_db)):
    """Generate a formatted Excel export of an MPD dataset."""
    if (r := _require_login(request)):
        return r
    ds = (await db.execute(select(MPDDataset).where(MPDDataset.id == dataset_id))).scalars().one_or_none()
    if not ds:
        raise HTTPException(404, "Dataset not found")
    tasks_res = await db.execute(
        select(MPDTask).where(MPDTask.dataset_id == dataset_id).order_by(MPDTask.row_index)
    )
    tasks = tasks_res.scalars().all()

    wb = Workbook()
    _DAE_FILL = PatternFill("solid", fgColor="C00000")
    _WHITE_BOLD = Font(color="FFFFFF", bold=True, size=10)
    _BOLD = Font(bold=True, size=10)
    _NORMAL = Font(size=10)
    _WRAP_TOP = Alignment(wrap_text=True, vertical="top")

    # ── Sheet 1: Info ───────────────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.merge_cells("A1:B1")
    c = ws_info["A1"]
    c.value = "MPD REFERENCE EXPORT — FOR REFERENCE ONLY"
    c.fill = _DAE_FILL
    c.font = Font(color="FFFFFF", bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_info.row_dimensions[1].height = 22

    info_rows = [
        ("Manufacturer", ds.manufacturer or ""),
        ("Model", ds.model or ""),
        ("Revision", ds.revision or ""),
        ("Tasks imported", str(len(tasks))),
        ("Export date", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Note", "This file is for reference only. Always verify against the current approved MPD."),
    ]
    for i, (k, v) in enumerate(info_rows, start=2):
        ws_info.cell(row=i, column=1, value=k).font = _BOLD
        cell = ws_info.cell(row=i, column=2, value=v)
        cell.font = _NORMAL
        cell.alignment = Alignment(wrap_text=True)
    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 60

    # ── Sheet 2: Tasks ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Tasks")

    all_headers = [
        "#", "MPD ITEM #", "TASK REFERENCE", "SECTION",
        "TITLE", "DESCRIPTION",
        "THRESHOLD", "INTERVAL", "EFFECTIVITY",
        "ZONE MH", "ACCESS MH", "PREP MH", "TOTAL MH", "SKILL",
    ]
    all_widths = [5, 14, 16, 10, 30, 40, 20, 20, 30, 8, 8, 8, 8, 14]

    def _to_lines(raw: str | None) -> str:
        """Convert comma/semicolon-separated values to one per line."""
        if not raw:
            return ""
        parts = [p.strip() for p in _re.split(r"[,;]\s*", raw) if p.strip()]
        return "\n".join(parts) if len(parts) > 1 else raw

    # Build all data rows first so we can detect empty columns
    data_rows: list[list] = []
    for seq, t in enumerate(tasks, start=1):
        try:
            mh_z = float(t.zone_mh) if t.zone_mh else 0.0
        except Exception:
            mh_z = 0.0
        try:
            mh_a = float(t.access_mh) if t.access_mh else 0.0
        except Exception:
            mh_a = 0.0
        try:
            mh_p = float(t.preparation_mh) if t.preparation_mh else 0.0
        except Exception:
            mh_p = 0.0
        mh_total = mh_z + mh_a + mh_p

        data_rows.append([
            seq,                          # sequential, unique, starts at 1
            t.mpd_item_number,
            t.task_reference,
            t.section,
            t.title,
            t.description,
            _to_lines(t.threshold_raw),
            _to_lines(t.interval_raw),
            _to_lines(t.applicability_raw),
            mh_z or None,
            mh_a or None,
            mh_p or None,
            mh_total or None,
            t.skill,
        ])

    # Determine which columns have at least one non-empty value (col index 0 = "#" always kept)
    def _is_empty_val(v) -> bool:
        if v is None:
            return True
        if isinstance(v, str):
            return v.strip() == ""
        if isinstance(v, float):
            return v == 0.0
        return False

    keep_cols = [
        ci for ci, h in enumerate(all_headers)
        if ci == 0 or any(not _is_empty_val(row[ci]) for row in data_rows)
    ]
    headers  = [all_headers[ci] for ci in keep_cols]
    widths   = [all_widths[ci]  for ci in keep_cols]

    for ci_out, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=ci_out, value=h)
        cell.fill = _DAE_FILL
        cell.font = _WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci_out)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for ri, row_vals_full in enumerate(data_rows, start=2):
        row_vals = [row_vals_full[ci] for ci in keep_cols]
        max_lines = 1
        for ci_out, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci_out, value=val)
            cell.font = _NORMAL
            cell.alignment = _WRAP_TOP
            if isinstance(val, str):
                max_lines = max(max_lines, val.count("\n") + 1)
        ws.row_dimensions[ri].height = max(14, min(max_lines * 13, 80))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_mfr = (ds.manufacturer or "").replace(" ", "_").replace("/", "-")
    safe_mdl = (ds.model or "").replace(" ", "_").replace("/", "-")
    safe_rev = (ds.revision or "").replace(" ", "_").replace("/", "-")
    filename = f"MPD_{safe_mfr}_{safe_mdl}_{safe_rev}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/mpd/{dataset_id}/delete")
async def mpd_dataset_delete(request: Request, dataset_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    if request.session.get("role") != "admin":
        raise HTTPException(403, "Admin only")

    # Block deletion if any project references this dataset
    in_use = (await db.execute(
        select(func.count()).select_from(Project).where(Project.mpd_dataset_id == dataset_id)
    )).scalar() or 0
    if in_use:
        return RedirectResponse(f"/mpd?error=in_use_{dataset_id}", status_code=303)

    ds = (await db.execute(select(MPDDataset).where(MPDDataset.id == dataset_id))).scalars().one_or_none()
    if ds:
        # Remove stored source file
        for ext in (".xlsx", ".xls"):
            fp = MPD_STORAGE / f"mpd_{dataset_id}{ext}"
            fp.unlink(missing_ok=True)
        await db.delete(ds)
        await db.commit()
    return RedirectResponse("/mpd", status_code=303)


@router.post("/mpd/{dataset_id}/toggle-superseded")
async def mpd_dataset_toggle_superseded(request: Request, dataset_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    if request.session.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    ds = (await db.execute(select(MPDDataset).where(MPDDataset.id == dataset_id))).scalars().one_or_none()
    if not ds:
        raise HTTPException(404, "Dataset not found")
    ds.is_superseded = not ds.is_superseded
    await db.commit()
    return RedirectResponse("/mpd", status_code=303)


@router.get("/mpd/{dataset_id}/tasks/{task_id}/edit", response_class=HTMLResponse)
async def mpd_task_edit_get(request: Request, dataset_id: int, task_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    t_res = await db.execute(select(MPDTask).where(MPDTask.id == task_id, MPDTask.dataset_id == dataset_id))
    task = t_res.scalars().one_or_none()
    if not task:
        raise HTTPException(404, "Task not found")
    ds_res = await db.execute(select(MPDDataset).where(MPDDataset.id == dataset_id))
    dataset = ds_res.scalars().one_or_none()
    return templates.TemplateResponse(request, "mpd_task_edit.html", {"task": task, "dataset": dataset})


@router.post("/mpd/{dataset_id}/tasks/{task_id}/edit")
async def mpd_task_edit_post(
    request: Request,
    dataset_id: int,
    task_id: int,
    mpd_item_number: str = Form(""),
    task_reference: str = Form(""),
    title: str = Form(""),
    description: str = Form(""),
    threshold_raw: str = Form(""),
    interval_raw: str = Form(""),
    applicability_raw: str = Form(""),
    section: str = Form(""),
    zone_mh: str = Form(""),
    access_mh: str = Form(""),
    preparation_mh: str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    if (r := _require_login(request)):
        return r
    t_res = await db.execute(select(MPDTask).where(MPDTask.id == task_id, MPDTask.dataset_id == dataset_id))
    task = t_res.scalars().one_or_none()
    if not task:
        raise HTTPException(404, "Task not found")
    task.mpd_item_number = mpd_item_number.strip() or None
    task.task_reference   = task_reference.strip()   or None
    task.title            = title.strip()            or None
    task.description      = description.strip()      or None
    task.threshold_raw    = threshold_raw.strip()    or None
    task.interval_raw     = interval_raw.strip()     or None
    task.applicability_raw = applicability_raw.strip() or None
    task.section          = section.strip()          or None
    task.zone_mh          = zone_mh.strip()          or None
    task.access_mh        = access_mh.strip()        or None
    task.preparation_mh   = preparation_mh.strip()   or None
    tokens = normalize_applicability_tokens(applicability_raw.strip() or None)
    task.applicability_tokens_normalized = ",".join(tokens) if tokens else None
    await db.commit()
    return RedirectResponse(f"/mpd/{dataset_id}", status_code=303)


@router.post("/mpd/{dataset_id}/tasks/{task_id}/delete")
async def mpd_task_delete(request: Request, dataset_id: int, task_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    t_res = await db.execute(select(MPDTask).where(MPDTask.id == task_id, MPDTask.dataset_id == dataset_id))
    task = t_res.scalars().one_or_none()
    if task:
        await db.delete(task)
        await db.commit()
    return RedirectResponse(f"/mpd/{dataset_id}", status_code=303)


@router.get("/report/{project_id}", response_class=HTMLResponse)
async def report_page(request: Request, project_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalars().one_or_none()
    if not project:
        raise HTTPException(404, "Project not found")
    from app.services.reporting import get_report_summary
    summary = await get_report_summary(db, project_id, project.mpd_dataset_id or 0)
    summary_json = json.dumps(summary)
    return templates.TemplateResponse(
        request, "report.html",
        {"project": project, "summary": summary, "summary_json": summary_json},
    )


# ── Effectivity crosscheck ────────────────────────────────────────────────────

@router.get("/projects/{project_id}/effectivity", response_class=HTMLResponse)
async def effectivity_view(request: Request, project_id: int, db: AsyncSession = Depends(get_db)):
    if (r := _require_login(request)):
        return r
    project = (await db.execute(select(Project).where(Project.id == project_id))).scalars().one_or_none()
    if not project:
        raise HTTPException(404)

    # Get all unique applicability tokens from the linked MPD dataset
    token_task_map: dict[str, list[dict]] = {}
    if project.mpd_dataset_id:
        tasks_res = await db.execute(
            select(MPDTask).where(
                MPDTask.dataset_id == project.mpd_dataset_id,
                MPDTask.applicability_tokens_normalized.isnot(None),
            )
        )
        def _safe_float(v):
            try:
                return float(v) if v else None
            except Exception:
                return None

        for t in tasks_res.scalars().all():
            raw = t.applicability_tokens_normalized or ""
            tokens = [tok.strip() for tok in raw.split(",") if tok.strip() and tok.strip().upper() != "ALL"]
            mh_parts = [_safe_float(t.zone_mh), _safe_float(t.access_mh), _safe_float(t.preparation_mh)]
            mh_total = sum(v for v in mh_parts if v) or None
            for tok in tokens:
                tok_up = tok.upper()
                if tok_up not in token_task_map:
                    token_task_map[tok_up] = []
                token_task_map[tok_up].append({
                    "id": t.id,
                    "item_no": t.mpd_item_number or "",
                    "ref": t.task_reference or "",
                    "section": t.section or "",
                    "title": t.title or "",
                    "description": t.description or "",
                    "threshold_raw": t.threshold_raw or "",
                    "interval_raw": t.interval_raw or "",
                    "applicability_raw": t.applicability_raw or "",
                    "skill": t.skill or "",
                    "mh_total": mh_total,
                })

    # Get current answers for this project
    answers_res = await db.execute(
        select(ProjectConditionAnswer).where(ProjectConditionAnswer.project_id == project_id)
    )
    answers = {a.condition_token: a for a in answers_res.scalars().all()}

    conditions = sorted(token_task_map.keys())
    condition_rows = []
    for tok in conditions:
        tasks_for_tok = token_task_map[tok]
        ans = answers.get(tok)
        condition_rows.append({
            "token": tok,
            "answer": ans.answer if ans else "TBC",
            "updated_at": ans.updated_at.strftime("%Y-%m-%d %H:%M") if ans and ans.updated_at else "",
            "task_count": len(tasks_for_tok),
            "tasks": tasks_for_tok,
        })

    return templates.TemplateResponse(request, "effectivity.html", {
        "project": project,
        "condition_rows": condition_rows,
        "project_id": project_id,
        "has_mpd": bool(project.mpd_dataset_id),
    })


@router.post("/projects/{project_id}/effectivity/answer")
async def effectivity_set_answer(
    request: Request,
    project_id: int,
    token: str = Form(...),
    answer: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    if (r := _require_login(request)):
        return r
    project = (await db.execute(select(Project).where(Project.id == project_id))).scalars().one_or_none()
    if not project:
        raise HTTPException(404)

    answer = answer.upper()
    if answer not in ("YES", "NO", "TBC"):
        raise HTTPException(422, "answer must be YES, NO, or TBC")

    username = request.session.get("username", "")

    existing = (await db.execute(
        select(ProjectConditionAnswer).where(
            ProjectConditionAnswer.project_id == project_id,
            ProjectConditionAnswer.condition_token == token,
        )
    )).scalars().one_or_none()

    old_answer = existing.answer if existing else None

    if existing:
        existing.answer = answer
        existing.source = "user"
        from datetime import datetime
        existing.updated_at = datetime.utcnow()
    else:
        db.add(ProjectConditionAnswer(
            project_id=project_id,
            condition_token=token,
            answer=answer,
            source="user",
        ))

    # Record history
    if old_answer != answer:
        db.add(ConditionAnswerHistory(
            project_id=project_id,
            condition_token=token,
            old_answer=old_answer,
            new_answer=answer,
            changed_by=username,
        ))

    await db.commit()
    return {"ok": True, "token": token, "answer": answer}
